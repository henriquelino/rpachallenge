
import logging
import sys
import openpyxl

import PythonUtils.chromedriver as chromedriver
import PythonUtils.log as log

from pathlib import Path
from PythonUtils.utils import APP_PATH

log.create_logger()
logger = logging.getLogger(__name__)

logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("selenium").setLevel(logging.WARNING)

class UserInfo:
    def __init__(self, first_name, last_name, company_name, role, address, email, phone):
        self.first_name = first_name
        self.last_name = last_name
        self.company_name = company_name
        self.role_in_company = role
        self.address = address
        self.email = email
        self.phone_number = phone
        return

def create_chrome():

    chrome_configs = {
        "driver_path" : APP_PATH / "chrome/chromedriver.exe",
        "id_path" : APP_PATH / "chrome/id.json",
    }

    chrome_options = {
        "arguments" : [
            "log-level=3",
            "no-first-run",
            # "incognito",
            "no-default-browser-check",
            "disable-infobars",
            "disable-blink-features",
            "disable-blink-features=AutomationControlled",
            f"user-data-dir={APP_PATH / 'chrome/data-dir'}",
        ],

        "experimental" : {
            "prefs" : {
                "download.prompt_for_download": False,
                "download.directory_upgrade" : True,
                "plugins.always_open_pdf_externally": True,
                "download.default_directory" : f"{ APP_PATH / 'chrome/downloads'}",
            },
            "excludeSwitches" : ['enable-automation', 'ignore-certificate-errors'],
            "useAutomationExtension" : False
        }
    }

    options = chromedriver.Options()

    for arg in chrome_options["arguments"]:
        options.add_argument(arg)

    for k, v in chrome_options["experimental"].items():
        options.add_experimental_option(k, v)



    driver = chromedriver.CustomChrome(**chrome_configs, options=options)
    # driver = chromedriver.CustomChrome(**chrome_configs)
    if not driver.begin():
        logger.critical(f"Falha ao criar chromedriver")
        return False
    
    # driver.rotate_user_agent()
    
    driver.set_navigator_to_undefined()

    driver.refresh()

    return driver

def download_excel(driver: chromedriver.CustomChrome):
    scripts = {
        "download_button" : 'document.querySelector("body > app-root > div.body.row1.scroll-y > app-rpa1 > div > div.instructions.col.s3.m3.l3.uiColorSecondary > div:nth-child(7) > a")',
    }

    r = driver.find_and_click_element(script=scripts["download_button"], timeout=30)
    if not r:
        logger.critical(f"Não consegui clicar no botão de download")
        return False

    r = driver.wait_all_downloads_end()
    if not r:
        logger.critical(f"Timeout aguardando os downloads acabarem")
        return False

    return True

def read_sheet(sheet_path: str) -> list[UserInfo]:

    wb_obj = openpyxl.load_workbook(sheet_path) 

    sheet = wb_obj.active
    
    #['First Name', 'Last Name ', 'Company Name', 'Role in Company', 'Address', 'Email', 'Phone Number']
    infos = []
    for row in sheet.iter_rows(max_col=sheet.max_column-1, values_only=True):
        if row[0] is None:
            break
        if row[0] == "First Name":
            continue

        infos.append(UserInfo(*row))
    
    return infos

def fill_rpa_form(driver: chromedriver.CustomChrome, infos: list[UserInfo]):

    logger.critical(f"Clicando em 'Start'")
    r = driver.find_and_click_element(script='document.querySelector("div > button")')
    if not r:
        logger.critical(f"Não encontrei o botão 'Start'")
        return False

    for info in infos:
        # 3 formas de preencher
        # a primeira usa um biblioteca externa, que valida o preenchimento
        # a segunda busca todos os campos separadamente com xpath e faz o preenchimento
        # a terceira busca os elementos pelo tag name
        
        # ~~~~~~~~~~~~~~~~~~~~~~~ 15s, valida o preenchimento ~~~~~~~~~~~~~~~~~~~~~~~
        # driver.find_and_fill_element(info.first_name,      by="xpath", selector='//input[@ng-reflect-name="labelFirstName"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.last_name,       by="xpath", selector='//input[@ng-reflect-name="labelLastName"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.address,         by="xpath", selector='//input[@ng-reflect-name="labelAddress"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.role_in_company, by="xpath", selector='//input[@ng-reflect-name="labelRole"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.email,           by="xpath", selector='//input[@ng-reflect-name="labelEmail"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.phone_number,    by="xpath", selector='//input[@ng-reflect-name="labelPhone"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_fill_element(info.company_name,    by="xpath", selector='//input[@ng-reflect-name="labelCompanyName"]', clear_before_fill=False, tab_after_fill=False)
        # driver.find_and_click_element(by="xpath", selector='//input[@type="submit"]')
        # continue

        # ~~~~~~~~~~~~~~~~~~~~~~~ 1~1.2s, não valida preenchimentos ~~~~~~~~~~~~~~~~~~~~~~~
        FirstName   = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelFirstName"]')
        LastName    = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelLastName"]')
        Address     = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelAddress"]')
        Role        = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelRole"]')
        Email       = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelEmail"]')
        Phone       = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelPhone"]')
        CompanyName = driver.find_element(by="xpath", value ='//input[@ng-reflect-name="labelCompanyName"]')
        Submit      = driver.find_element(by="xpath", value='//input[@type="submit"]')

        driver.execute_script(f"arguments[0].value='{info.first_name}';",       FirstName)
        driver.execute_script(f"arguments[0].value='{info.last_name}';",        LastName)
        driver.execute_script(f"arguments[0].value='{info.address}';",          Address)
        driver.execute_script(f"arguments[0].value='{info.role_in_company}';",  Role)
        driver.execute_script(f"arguments[0].value='{info.email}';",            Email)
        driver.execute_script(f"arguments[0].value='{info.phone_number}';",     Phone)
        driver.execute_script(f"arguments[0].value='{info.company_name}';",     CompanyName)
        driver.execute_script(f"arguments[0].click();", Submit)
        continue

        # ~~~~~~~~~~~~~~~~~~~~~~~ 1.5s~, não valida preenchimento ~~~~~~~~~~~~~~~~~~~~~~~
        element = driver.find_element("tag name", "form")
        inputs : list[chromedriver.WebElement] = element.find_elements("tag name", "input")
        labels : list[chromedriver.WebElement] = element.find_elements("tag name", "label")
        for input, label in zip(inputs, labels):
            label = label.text.lower().replace(" ", "_")
            driver.execute_script(f"arguments[0].value='{info.__getattribute__(label)}';", input)
        driver.execute_script(f"arguments[0].click();", inputs[7])

        return
    
    return True

def main():
    
    driver = create_chrome()

    # --------------------
    r = driver.open_url("https://www.rpachallenge.com")
    if not r:
        sys.exit(f"Não consegui abrir o site do RPA Challenge")

    # --------------------
    r = download_excel(driver)
    if not r:
        sys.exit(f"Não consegui baixar a planilha")

    # --------------------
    # Busca a planilha que foi baixada anteriormente
    if driver.options: # na pasta do chrome
        download_folder = Path(driver.options.to_capabilities()["goog:chromeOptions"]["prefs"]["download.default_directory"])
    else: # ou na pasta de downloads do usuário
        download_folder = Path('~/Downloads').expanduser()
    
    # procura por todos arquivos .xlsx
    sheets = list(download_folder.glob("*.xlsx"))
    if not sheets:
        sys.exit("Planilha não foi baixada")
    
    # e escolhe o último arquivo
    last_sheet = max(sheets, key=lambda f: f.stat().st_ctime)
    
    # --------------------
    infos = read_sheet(last_sheet)

    r = fill_rpa_form(driver, infos)
    if not r:
        sys.exit("Falha no preenchimento do formulário")
    
    return

if __name__ == '__main__':
    main()