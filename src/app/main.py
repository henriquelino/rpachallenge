import logging
import sys
import time
from pathlib import Path
from shutil import rmtree

import definitions
import openpyxl
from log import setup_logging
from loguru import logger
from models import UserInfo
from selenium.webdriver.remote.webelement import WebElement
from selenium_tkit.chrome import ChromeOptions, CreateChrome
from selenium_tkit.custom_webdriver import CustomWebDriver

output_png = definitions.BASE_DIR / 'metrics.png'

setup_logging(definitions.BASE_DIR)
logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("selenium").setLevel(logging.WARNING)


def create_chrome():
    """This is an example function that you can copy and customize for each project"""

    chrome_configs = {
        "driver_path": definitions.BASE_DIR / "chromedriver",
    }

    chrome_options = {
        "arguments": [
            "log-level=3",
            "no-first-run",
            # "incognito",
            "no-default-browser-check",
            "disable-infobars",
            "disable-blink-features",
            "disable-blink-features=AutomationControlled",
            rf"user-data-dir={definitions.BASE_DIR}\chromedriver\data-dir",
        ],
        "experimental": {
            "prefs": {
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True,
                "download.default_directory": str(definitions.BASE_DIR / "chromedriver/downloads"),
            },
            "excludeSwitches": ["enable-automation", "ignore-certificate-errors"],
            "useAutomationExtension": False,
        },
        "extensions": [
            # r"path_to\extensions" # --> a folder will use all .crx of folder
            # r"path_to\extensions\uBlock-Origin.crx", # --> only a single extension
        ],
    }  # yapf: disable

    options = ChromeOptions()

    # --------------------------------------------------
    # add arguments
    for arg in chrome_options["arguments"]:
        options.add_argument(arg)

    # --------------------------------------------------
    # add experimental options
    for k, v in chrome_options["experimental"].items():
        options.add_experimental_option(k, v)

    # --------------------------------------------------
    # if extension isn't a list or a tuple, fixes to a list
    if chrome_options["extensions"] and not isinstance(chrome_options["extensions"], (list, tuple)):
        chrome_options["extensions"] = list(chrome_options["extensions"])

    # add extensions
    all_extensions = list()
    for ext in chrome_options["extensions"]:
        ext = Path(ext)
        if ext.is_dir():
            for e in ext.glob("*.*"):
                all_extensions.append(str(e))
        else:
            all_extensions.append(str(ext))

    for ext in all_extensions:
        options.add_extension(ext)

    # --------------------------------------------------

    driver = CreateChrome(**chrome_configs, options=options)
    if not driver.begin():
        logger.info("Something went wrong creating a chrome instance. Check logs for details.")
        return False

    # --------------------------------------------------

    # driver.rotate_user_agent()

    driver.set_navigator_to_undefined()

    driver._is_remote = False

    return driver


def cleanup():
    if output_png.exists():
        output_png.unlink()

    if (download_dir := (definitions.BASE_DIR / 'chromedriver' / 'downloads')).exists():
        rmtree(download_dir)


def download_excel(driver: CustomWebDriver):
    scripts = {
        "download_button": {
            'by': 'xpath',
            'selector': '//a[contains(text(), "Download Excel")]',
        }
    }
    logger.debug("Clicando para baixar a planilha")
    r = driver.find_and_click_element(**scripts["download_button"])
    if not r:
        logger.info("Não consegui clicar no botão de download")
        return False

    time.sleep(1)  # wait download begins?

    logger.debug("Aguardando download finalizar")
    r = driver.wait_all_downloads_end()
    if not r:
        logger.info("Timeout aguardando os downloads acabarem")
        return False

    return True


def read_sheet(sheet_path: str) -> list[UserInfo]:

    wb_obj = openpyxl.load_workbook(sheet_path)

    sheet = wb_obj.active

    # ['First Name', 'Last Name ', 'Company Name', 'Role in Company', 'Address', 'Email', 'Phone Number']
    infos = []
    for row in sheet.iter_rows(max_col=sheet.max_column - 1, values_only=True):
        if row[0] is None:
            break
        if row[0] == "First Name":
            continue

        infos.append(UserInfo(*row))

    return infos


def fill_rpa_form(driver: CustomWebDriver, infos: list[UserInfo]):

    logger.info("Clicando em 'Start'")
    r = driver.find_and_click_element(by='xpath', selector='//div/button[contains(text(), "Start")]')
    if not r:
        logger.info("Não encontrei o botão 'Start'")
        return False

    # 3 formas de preencher
    # 1 -> valida o preenchimento, um pouco mais demorado
    # 2 -> a segunda busca todos os campos e preenchendo um a um
    # 3 -> terceira busca os labels e inputs pelo tag name, depois preenche
    method = 2
    for info in infos:
        if method == 1:
            # ~~~~~~~~~~~~~~~~~~~~~~~ 15s, valida o preenchimento ~~~~~~~~~~~~~~~~~~~~~~~
            # find_and_fill_element e find_and_click_element são métodos do selenium-tkit
            driver.find_and_fill_element(info.first_name, by="xpath", selector='//input[@ng-reflect-name="labelFirstName"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(info.last_name, by="xpath", selector='//input[@ng-reflect-name="labelLastName"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(info.address, by="xpath", selector='//input[@ng-reflect-name="labelAddress"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(info.role_in_company, by="xpath", selector='//input[@ng-reflect-name="labelRole"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(info.email, by="xpath", selector='//input[@ng-reflect-name="labelEmail"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(str(info.phone_number), by="xpath", selector='//input[@ng-reflect-name="labelPhone"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_fill_element(info.company_name, by="xpath", selector='//input[@ng-reflect-name="labelCompanyName"]', clear_before_fill=False, tab_after_fill=False)
            driver.find_and_click_element(by="xpath", selector='//input[@type="submit"]')

        elif method == 2:
            # ~~~~~~~~~~~~~~~~~~~~~~~ 1~1.2s, não valida preenchimentos ~~~~~~~~~~~~~~~~~~~~~~~
            # find_element e execute_script são os métodos padrões do Selenium
            first_name = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelFirstName"]')
            driver.execute_script(f"arguments[0].value='{info.first_name}';", first_name)

            last_name = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelLastName"]')
            driver.execute_script(f"arguments[0].value='{info.last_name}';", last_name)

            address = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelAddress"]')
            driver.execute_script(f"arguments[0].value='{info.address}';", address)

            role = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelRole"]')
            driver.execute_script(f"arguments[0].value='{info.role_in_company}';", role)

            email = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelEmail"]')
            driver.execute_script(f"arguments[0].value='{info.email}';", email)

            phone = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelPhone"]')
            driver.execute_script(f"arguments[0].value='{info.phone_number}';", phone)

            company_name = driver.find_element(by="xpath", value='//input[@ng-reflect-name="labelCompanyName"]')
            driver.execute_script(f"arguments[0].value='{info.company_name}';", company_name)

            Submit = driver.find_element(by="xpath", value='//input[@type="submit"]')
            driver.execute_script("arguments[0].click();", Submit)

        elif method == 3:
            # ~~~~~~~~~~~~~~~~~~~~~~~ 1.5s~, não valida preenchimento ~~~~~~~~~~~~~~~~~~~~~~~
            element = driver.find_element("tag name", "form")
            inputs: list[WebElement] = element.find_elements("tag name", "input")
            labels: list[WebElement] = element.find_elements("tag name", "label")
            for input, label in zip(inputs, labels):
                label = label.text.lower().replace(" ", "_")
                driver.execute_script(f"arguments[0].value='{info.__getattribute__(label)}';", input)
            driver.execute_script("arguments[0].click();", inputs[7])

        else:
            raise Exception("Unknow fill method")

    return True


def main():
    cleanup()

    driver = create_chrome()

    # --------------------
    r = driver.open_url("https://www.rpachallenge.com")
    if not r:
        sys.exit("Não consegui abrir o site do RPA Challenge")

    # --------------------
    r = download_excel(driver)
    if not r:
        sys.exit("Não consegui baixar a planilha")

    # --------------------
    # Busca a planilha que foi baixada anteriormente
    if driver.options:  # na pasta do chrome
        download_folder = Path(driver.options.to_capabilities()["goog:chromeOptions"]["prefs"]["download.default_directory"])
    else:  # ou na pasta de downloads do usuário
        download_folder = Path('~/Downloads').expanduser()

    # procura por todos arquivos .xlsx
    sheets = list(download_folder.glob("*.xlsx"))
    if not sheets:
        sys.exit("Planilha não foi baixada")

    # e escolhe o último arquivo
    last_sheet = max(sheets, key=lambda f: f.stat().st_ctime)

    # --------------------

    infos = read_sheet(last_sheet)

    r = fill_rpa_form(driver, infos)
    if not r:
        sys.exit("Falha no preenchimento do formulário")
    driver.save_screenshot(output_png)
    logger.critical(f"Screenshot salva em: '{output_png}'")
    return


if __name__ == '__main__':
    main()
